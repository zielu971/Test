import subprocess

import pyperclip as clipboard
import os
from random import randint
import openpyxl
import pandas as pd
from selenium import webdriver
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QLineEdit, QApplication
from time import sleep
from threading import *

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

ver = "Beta v2.6"


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(631, 875)
        MainWindow.setMinimumSize(QtCore.QSize(631, 875))
        MainWindow.setMaximumSize(QtCore.QSize(631, 875))
        MainWindow.setStyleSheet("")
        MainWindow.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        MainWindow.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(10, 10, 611, 841))
        self.widget.setObjectName("widget")
        self.BG = QtWidgets.QLabel(self.widget)
        self.BG.setGeometry(QtCore.QRect(10, 0, 601, 841))
        self.BG.setMinimumSize(QtCore.QSize(601, 841))
        self.BG.setStyleSheet("background-color: rgb(40,36,46);\n"
"border-radius:50px;\n"
"")
        self.BG.setText("")
        self.BG.setObjectName("BG")
        self.LoginIGmainLabel = QtWidgets.QLabel(self.widget)
        self.LoginIGmainLabel.setGeometry(QtCore.QRect(250, 60, 121, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LoginIGmainLabel.setFont(font)
        self.LoginIGmainLabel.setStyleSheet("color: rgb(255, 255, 255);")
        self.LoginIGmainLabel.setObjectName("LoginIGmainLabel")
        self.label = QtWidgets.QLabel(self.widget)
        self.label.setGeometry(QtCore.QRect(20, 60, 580, 130))
        self.label.setStyleSheet("background-color: rgb(53,48,61);\n"
"border-radius:50px;")
        self.label.setText("")
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.widget)
        self.label_2.setGeometry(QtCore.QRect(126, 90, 51, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.widget)
        self.label_3.setGeometry(QtCore.QRect(126, 120, 51, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_3.setObjectName("label_3")
        self.loginEntry = QtWidgets.QLineEdit(self.widget)
        self.loginEntry.setGeometry(QtCore.QRect(180, 90, 161, 20))
        self.loginEntry.setStyleSheet("background-color: rgb(79,74,84);\n"
"color: rgb(255, 255, 255);\n"
"border:none;\n"
"border-radius:10px;\n"
"\n"
"")
        self.loginEntry.setEchoMode(QtWidgets.QLineEdit.Normal)
        self.loginEntry.setPlaceholderText("")
        self.loginEntry.setObjectName("loginEntry")
        self.passwordEntry = QtWidgets.QLineEdit(self.widget)
        self.passwordEntry.setGeometry(QtCore.QRect(180, 120, 161, 20))
        self.passwordEntry.setStyleSheet("background-color: rgb(79,74,84);\n"
"color: rgb(255, 255, 255);\n"
"border:none;\n"
"border-radius:10px;\n"
"\n"
"")
        self.passwordEntry.setText("")
        self.passwordEntry.setEchoMode(QtWidgets.QLineEdit.Password)
        self.passwordEntry.setObjectName("passwordEntry")
        self.saveButton = QtWidgets.QToolButton(self.widget)
        self.saveButton.setGeometry(QtCore.QRect(180, 150, 81, 23))
        self.saveButton.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.saveButton.setObjectName("saveButton")
        self.loadButton = QtWidgets.QToolButton(self.widget)
        self.loadButton.setGeometry(QtCore.QRect(270, 150, 81, 23))
        self.loadButton.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.loadButton.setObjectName("loadButton")
        self.delButton = QtWidgets.QToolButton(self.widget)
        self.delButton.setGeometry(QtCore.QRect(360, 150, 81, 23))
        self.delButton.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.delButton.setObjectName("delButton")
        self.label_4 = QtWidgets.QLabel(self.widget)
        self.label_4.setGeometry(QtCore.QRect(435, 86, 51, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.widget)
        self.label_5.setGeometry(QtCore.QRect(435, 116, 51, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_5.setFont(font)
        self.label_5.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_5.setObjectName("label_5")
        self.FirefoxButton = QtWidgets.QRadioButton(self.widget)
        self.FirefoxButton.setGeometry(QtCore.QRect(490, 90, 16, 17))
        self.FirefoxButton.setText("")
        self.FirefoxButton.setObjectName("FirefoxButton")
        self.Chromebutton = QtWidgets.QRadioButton(self.widget)
        self.Chromebutton.setGeometry(QtCore.QRect(490, 120, 21, 16))
        self.Chromebutton.setText("")
        self.Chromebutton.setChecked(True)
        self.Chromebutton.setObjectName("Chromebutton")
        self.label_6 = QtWidgets.QLabel(self.widget)
        self.label_6.setGeometry(QtCore.QRect(20, 200, 290, 130))
        self.label_6.setStyleSheet("background-color: rgb(53,48,61);\n"
"border-radius:50px;")
        self.label_6.setText("")
        self.label_6.setObjectName("label_6")
        self.LoginIGmainLabel_2 = QtWidgets.QLabel(self.widget)
        self.LoginIGmainLabel_2.setGeometry(QtCore.QRect(70, 210, 191, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LoginIGmainLabel_2.setFont(font)
        self.LoginIGmainLabel_2.setStyleSheet("color: rgb(255, 255, 255);")
        self.LoginIGmainLabel_2.setObjectName("LoginIGmainLabel_2")
        self.IloscKontEntry = QtWidgets.QLineEdit(self.widget)
        self.IloscKontEntry.setGeometry(QtCore.QRect(130, 270, 170, 20))
        self.IloscKontEntry.setStyleSheet("background-color: rgb(79,74,84);\n"
"color: rgb(255, 255, 255);\n"
"border:none;\n"
"border-radius:10px;\n"
"\n"
"")
        self.IloscKontEntry.setObjectName("IloscKontEntry")
        self.NazwaKontaEntry = QtWidgets.QLineEdit(self.widget)
        self.NazwaKontaEntry.setGeometry(QtCore.QRect(130, 240, 170, 20))
        self.NazwaKontaEntry.setStyleSheet("background-color: rgb(79,74,84);\n"
"color: rgb(255, 255, 255);\n"
"border: none;\n"
"border-radius:10px;\n"
"\n"
"")
        self.NazwaKontaEntry.setPlaceholderText("")
        self.NazwaKontaEntry.setObjectName("NazwaKontaEntry")
        self.label_7 = QtWidgets.QLabel(self.widget)
        self.label_7.setGeometry(QtCore.QRect(30, 240, 81, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_7.setFont(font)
        self.label_7.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(self.widget)
        self.label_8.setGeometry(QtCore.QRect(30, 270, 81, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_8.setFont(font)
        self.label_8.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_8.setObjectName("label_8")
        self.FollowersButton = QtWidgets.QToolButton(self.widget)
        self.FollowersButton.setGeometry(QtCore.QRect(95, 300, 141, 23))
        palette = QtGui.QPalette()
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
        brush.setStyle(QtCore.Qt.NoBrush)
        palette.setBrush(QtGui.QPalette.Active, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
        brush.setStyle(QtCore.Qt.NoBrush)
        palette.setBrush(QtGui.QPalette.Inactive, QtGui.QPalette.PlaceholderText, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.WindowText, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Button, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Text, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.ButtonText, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Base, brush)
        brush = QtGui.QBrush(QtGui.QColor(79, 74, 84))
        brush.setStyle(QtCore.Qt.SolidPattern)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.Window, brush)
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255, 128))
        brush.setStyle(QtCore.Qt.NoBrush)
        palette.setBrush(QtGui.QPalette.Disabled, QtGui.QPalette.PlaceholderText, brush)
        self.FollowersButton.setPalette(palette)
        self.FollowersButton.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.FollowersButton.setObjectName("FollowersButton")
        self.label_11 = QtWidgets.QLabel(self.widget)
        self.label_11.setGeometry(QtCore.QRect(310, 200, 290, 131))
        self.label_11.setStyleSheet("background-color: rgb(53,48,61);\n"
"border-radius:50px;")
        self.label_11.setText("")
        self.label_11.setObjectName("label_11")
        self.NazwaHaszEntry = QtWidgets.QLineEdit(self.widget)
        self.NazwaHaszEntry.setGeometry(QtCore.QRect(400, 240, 170, 20))
        self.NazwaHaszEntry.setStyleSheet("background-color: rgb(79,74,84);\n"
"color: rgb(255, 255, 255);\n"
"border:none;\n"
"border-radius:10px;\n"
"\n"
"")
        self.NazwaHaszEntry.setPlaceholderText("")
        self.NazwaHaszEntry.setObjectName("NazwaHaszEntry")
        self.HaszButton = QtWidgets.QToolButton(self.widget)
        self.HaszButton.setGeometry(QtCore.QRect(390, 300, 141, 23))
        self.HaszButton.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.HaszButton.setObjectName("HaszButton")
        self.LoginIGmainLabel_3 = QtWidgets.QLabel(self.widget)
        self.LoginIGmainLabel_3.setGeometry(QtCore.QRect(395, 210, 131, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LoginIGmainLabel_3.setFont(font)
        self.LoginIGmainLabel_3.setStyleSheet("color: rgb(255, 255, 255);")
        self.LoginIGmainLabel_3.setObjectName("LoginIGmainLabel_3")
        self.label_12 = QtWidgets.QLabel(self.widget)
        self.label_12.setGeometry(QtCore.QRect(320, 240, 71, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_12.setFont(font)
        self.label_12.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.label_13 = QtWidgets.QLabel(self.widget)
        self.label_13.setGeometry(QtCore.QRect(320, 270, 81, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_13.setFont(font)
        self.label_13.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_13.setObjectName("label_13")
        self.IloscHaszKont = QtWidgets.QLineEdit(self.widget)
        self.IloscHaszKont.setGeometry(QtCore.QRect(400, 270, 170, 20))
        self.IloscHaszKont.setStyleSheet("background-color: rgb(79,74,84);\n"
"color: rgb(255, 255, 255);\n"
"border:none;\n"
"border-radius:10px;\n"
"\n"
"")
        self.IloscHaszKont.setPlaceholderText("")
        self.IloscHaszKont.setObjectName("IloscHaszKont")
        self.label_15 = QtWidgets.QLabel(self.widget)
        self.label_15.setGeometry(QtCore.QRect(20, 340, 290, 161))
        self.label_15.setStyleSheet("background-color: rgb(53,48,61);\n"
"border-radius:50px;")
        self.label_15.setText("")
        self.label_15.setObjectName("label_15")
        self.LoginIGmainLabel_4 = QtWidgets.QLabel(self.widget)
        self.LoginIGmainLabel_4.setGeometry(QtCore.QRect(100, 350, 131, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LoginIGmainLabel_4.setFont(font)
        self.LoginIGmainLabel_4.setStyleSheet("color: rgb(255, 255, 255);")
        self.LoginIGmainLabel_4.setObjectName("LoginIGmainLabel_4")
        self.ComentCheckBox = QtWidgets.QCheckBox(self.widget)
        self.ComentCheckBox.setGeometry(QtCore.QRect(190, 380, 16, 17))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.ComentCheckBox.setFont(font)
        self.ComentCheckBox.setStyleSheet("color: rgb(255, 255, 255);border-radus:5px;")
        self.ComentCheckBox.setText("")
        self.ComentCheckBox.setIconSize(QtCore.QSize(20, 20))
        self.ComentCheckBox.setChecked(True)
        self.ComentCheckBox.setTristate(False)
        self.ComentCheckBox.setObjectName("ComentCheckBox")
        self.LikeCheckBox = QtWidgets.QCheckBox(self.widget)
        self.LikeCheckBox.setGeometry(QtCore.QRect(190, 410, 16, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.LikeCheckBox.setFont(font)
        self.LikeCheckBox.setStyleSheet("color: rgb(255, 255, 255);")
        self.LikeCheckBox.setText("")
        self.LikeCheckBox.setIconSize(QtCore.QSize(20, 20))
        self.LikeCheckBox.setChecked(True)
        self.LikeCheckBox.setObjectName("LikeCheckBox")
        self.label_16 = QtWidgets.QLabel(self.widget)
        self.label_16.setGeometry(QtCore.QRect(65, 380, 101, 20))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_16.setFont(font)
        self.label_16.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_16.setObjectName("label_16")
        self.label_17 = QtWidgets.QLabel(self.widget)
        self.label_17.setGeometry(QtCore.QRect(65, 410, 101, 21))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_17.setFont(font)
        self.label_17.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_17.setObjectName("label_17")
        self.FollowButton = QtWidgets.QToolButton(self.widget)
        self.FollowButton.setGeometry(QtCore.QRect(90, 460, 141, 23))
        self.FollowButton.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.FollowButton.setObjectName("FollowButton")
        self.label_18 = QtWidgets.QLabel(self.widget)
        self.label_18.setGeometry(QtCore.QRect(310, 340, 290, 161))
        self.label_18.setStyleSheet("background-color: rgb(53,48,61);\n"
"border-radius:50px;")
        self.label_18.setText("")
        self.label_18.setObjectName("label_18")
        self.LoginIGmainLabel_5 = QtWidgets.QLabel(self.widget)
        self.LoginIGmainLabel_5.setGeometry(QtCore.QRect(400, 350, 121, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LoginIGmainLabel_5.setFont(font)
        self.LoginIGmainLabel_5.setStyleSheet("color: rgb(255, 255, 255);")
        self.LoginIGmainLabel_5.setObjectName("LoginIGmainLabel_5")
        self.Message_Button = QtWidgets.QToolButton(self.widget)
        self.Message_Button.setGeometry(QtCore.QRect(390, 380, 141, 23))
        self.Message_Button.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.Message_Button.setObjectName("Message_Button")
        self.label_9 = QtWidgets.QLabel(self.widget)
        self.label_9.setGeometry(QtCore.QRect(20, 510, 580, 294))
        self.label_9.setStyleSheet("background-color: rgb(53,48,61);\n"
"border-radius:50px;")
        self.label_9.setText("")
        self.label_9.setObjectName("label_9")
        self.LoginIGmainLabel_6 = QtWidgets.QLabel(self.widget)
        self.LoginIGmainLabel_6.setGeometry(QtCore.QRect(265, 520, 91, 16))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LoginIGmainLabel_6.setFont(font)
        self.LoginIGmainLabel_6.setStyleSheet("color: rgb(255, 255, 255);")
        self.LoginIGmainLabel_6.setObjectName("LoginIGmainLabel_6")
        self.label_19 = QtWidgets.QLabel(self.widget)
        self.label_19.setGeometry(QtCore.QRect(30, 550, 81, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_19.setFont(font)
        self.label_19.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_19.setObjectName("label_19")
        self.label_10 = QtWidgets.QLabel(self.widget)
        self.label_10.setGeometry(QtCore.QRect(240, 0, 141, 51))
        self.label_10.setStyleSheet("border-image: url(:/logo/logo white bez tla.png);")
        self.label_10.setText("")
        self.label_10.setObjectName("label_10")
        self.label_20 = QtWidgets.QLabel(self.widget)
        self.label_20.setGeometry(QtCore.QRect(120, 550, 91, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_20.setFont(font)
        self.label_20.setStyleSheet("\n"
"color: rgb(123, 17, 58);")
        self.label_20.setAlignment(QtCore.Qt.AlignCenter)
        self.label_20.setObjectName("label_20")
        self.CommentEdit = QtWidgets.QToolButton(self.widget)
        self.CommentEdit.setGeometry(QtCore.QRect(390, 771, 141, 23))
        self.CommentEdit.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.CommentEdit.setObjectName("CommentEdit")
        self.MessageEdit = QtWidgets.QToolButton(self.widget)
        self.MessageEdit.setGeometry(QtCore.QRect(240, 771, 141, 23))
        self.MessageEdit.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.MessageEdit.setObjectName("MessageEdit")
        self.Website = QtWidgets.QToolButton(self.widget)
        self.Website.setGeometry(QtCore.QRect(90, 771, 141, 23))
        self.Website.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.Website.setObjectName("Website")
        self.label_21 = QtWidgets.QLabel(self.widget)
        self.label_21.setGeometry(QtCore.QRect(80, 810, 461, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_21.setFont(font)
        self.label_21.setStyleSheet("color: rgb(179, 179, 179);")
        self.label_21.setAlignment(QtCore.Qt.AlignCenter)
        self.label_21.setObjectName("label_21")
        self.label_22 = QtWidgets.QLabel(self.widget)
        self.label_22.setGeometry(QtCore.QRect(305, 550, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_22.setFont(font)
        self.label_22.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_22.setObjectName("label_22")
        self.timerLabel = QtWidgets.QLabel(self.widget)
        self.timerLabel.setGeometry(QtCore.QRect(495, 550, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.timerLabel.setFont(font)
        self.timerLabel.setStyleSheet("color: rgb(255, 255, 255);")
        self.timerLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.timerLabel.setObjectName("timerLabel")
        self.label_24 = QtWidgets.QLabel(self.widget)
        self.label_24.setGeometry(QtCore.QRect(305, 580, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_24.setFont(font)
        self.label_24.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_24.setObjectName("label_24")
        self.followCountLabel = QtWidgets.QLabel(self.widget)
        self.followCountLabel.setGeometry(QtCore.QRect(495, 580, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.followCountLabel.setFont(font)
        self.followCountLabel.setStyleSheet("color: rgb(255, 255, 255);")
        self.followCountLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.followCountLabel.setObjectName("followCountLabel")
        self.label_26 = QtWidgets.QLabel(self.widget)
        self.label_26.setGeometry(QtCore.QRect(305, 670, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_26.setFont(font)
        self.label_26.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_26.setObjectName("label_26")
        self.DownloadedLabel = QtWidgets.QLabel(self.widget)
        self.DownloadedLabel.setGeometry(QtCore.QRect(495, 670, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.DownloadedLabel.setFont(font)
        self.DownloadedLabel.setStyleSheet("color: rgb(255, 255, 255);")
        self.DownloadedLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.DownloadedLabel.setObjectName("DownloadedLabel")
        self.label_27 = QtWidgets.QLabel(self.widget)
        self.label_27.setGeometry(QtCore.QRect(305, 700, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_27.setFont(font)
        self.label_27.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_27.setObjectName("label_27")
        self.SendedMessage = QtWidgets.QLabel(self.widget)
        self.SendedMessage.setGeometry(QtCore.QRect(495, 700, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.SendedMessage.setFont(font)
        self.SendedMessage.setStyleSheet("color: rgb(255, 255, 255);")
        self.SendedMessage.setAlignment(QtCore.Qt.AlignCenter)
        self.SendedMessage.setObjectName("SendedMessage")
        self.label_25 = QtWidgets.QLabel(self.widget)
        self.label_25.setGeometry(QtCore.QRect(305, 640, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_25.setFont(font)
        self.label_25.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_25.setObjectName("label_25")
        self.label_28 = QtWidgets.QLabel(self.widget)
        self.label_28.setGeometry(QtCore.QRect(305, 610, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_28.setFont(font)
        self.label_28.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_28.setObjectName("label_28")
        self.LikedPhotoLabel = QtWidgets.QLabel(self.widget)
        self.LikedPhotoLabel.setGeometry(QtCore.QRect(495, 610, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LikedPhotoLabel.setFont(font)
        self.LikedPhotoLabel.setStyleSheet("color: rgb(255, 255, 255);")
        self.LikedPhotoLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.LikedPhotoLabel.setObjectName("LikedPhotoLabel")
        self.ComentLabel = QtWidgets.QLabel(self.widget)
        self.ComentLabel.setGeometry(QtCore.QRect(495, 640, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.ComentLabel.setFont(font)
        self.ComentLabel.setStyleSheet("color: rgb(255, 255, 255);")
        self.ComentLabel.setAlignment(QtCore.Qt.AlignCenter)
        self.ComentLabel.setObjectName("ComentLabel")
        self.CloseButton = QtWidgets.QToolButton(self.widget)
        self.CloseButton.setGeometry(QtCore.QRect(550, 20, 16, 16))
        font = QtGui.QFont()
        font.setFamily("Montserrat")
        font.setPointSize(7)
        font.setBold(False)
        font.setWeight(50)
        self.CloseButton.setFont(font)
        self.CloseButton.setToolTipDuration(1)
        self.CloseButton.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.CloseButton.setStyleSheet("background-color: rgb(255, 96, 92);\n"
"image: url(:/ex/close.png);\n"
"border-radius:8px;\n"
"")
        self.CloseButton.setText("")
        self.CloseButton.setIconSize(QtCore.QSize(25, 25))
        self.CloseButton.setToolButtonStyle(QtCore.Qt.ToolButtonIconOnly)
        self.CloseButton.setAutoRaise(False)
        self.CloseButton.setObjectName("CloseButton")
        self.minimaliseButton = QtWidgets.QToolButton(self.widget)
        self.minimaliseButton.setGeometry(QtCore.QRect(530, 20, 16, 16))
        self.minimaliseButton.setStyleSheet("background-color: rgb(255, 189, 68);\n"
"image: url(:/min/minimize.png);\n"
"border-radius:8px;\n"
"")
        self.minimaliseButton.setText("")
        self.minimaliseButton.setIconSize(QtCore.QSize(25, 25))
        self.minimaliseButton.setObjectName("minimaliseButton")
        self.label_14 = QtWidgets.QLabel(self.widget)
        self.label_14.setGeometry(QtCore.QRect(70, 20, 21, 16))
        self.label_14.setStyleSheet("color:rgb(255, 255, 255);")
        self.label_14.setObjectName("label_14")
        self.label_29 = QtWidgets.QLabel(self.widget)
        self.label_29.setGeometry(QtCore.QRect(30, 580, 81, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_29.setFont(font)
        self.label_29.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_29.setObjectName("label_29")
        self.WhoIsLogin = QtWidgets.QLabel(self.widget)
        self.WhoIsLogin.setGeometry(QtCore.QRect(120, 580, 81, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.WhoIsLogin.setFont(font)
        self.WhoIsLogin.setStyleSheet("color: rgb(255, 255, 255);")
        self.WhoIsLogin.setAlignment(QtCore.Qt.AlignCenter)
        self.WhoIsLogin.setObjectName("WhoIsLogin")
        self.label_30 = QtWidgets.QLabel(self.widget)
        self.label_30.setGeometry(QtCore.QRect(30, 610, 81, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_30.setFont(font)
        self.label_30.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_30.setObjectName("label_30")
        self.HowMuchDays = QtWidgets.QLabel(self.widget)
        self.HowMuchDays.setGeometry(QtCore.QRect(120, 610, 81, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.HowMuchDays.setFont(font)
        self.HowMuchDays.setStyleSheet("color: rgb(255, 255, 255);")
        self.HowMuchDays.setAlignment(QtCore.Qt.AlignCenter)
        self.HowMuchDays.setObjectName("HowMuchDays")
        self.Ver = QtWidgets.QLabel(self.widget)
        self.Ver.setGeometry(QtCore.QRect(90, 20, 61, 16))
        self.Ver.setStyleSheet("color:rgb(255, 255, 255);")
        self.Ver.setAlignment(QtCore.Qt.AlignCenter)
        self.Ver.setObjectName("Ver")
        self.LoginIGmainLabel_7 = QtWidgets.QLabel(self.widget)
        self.LoginIGmainLabel_7.setGeometry(QtCore.QRect(430, 410, 61, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.LoginIGmainLabel_7.setFont(font)
        self.LoginIGmainLabel_7.setStyleSheet("color: rgb(255, 255, 255);")
        self.LoginIGmainLabel_7.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.LoginIGmainLabel_7.setFrameShadow(QtWidgets.QFrame.Plain)
        self.LoginIGmainLabel_7.setObjectName("LoginIGmainLabel_7")
        self.DodajKonto = QtWidgets.QToolButton(self.widget)
        self.DodajKonto.setGeometry(QtCore.QRect(390, 430, 141, 23))
        self.DodajKonto.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.DodajKonto.setObjectName("DodajKonto")
        self.UnfollowButton = QtWidgets.QToolButton(self.widget)
        self.UnfollowButton.setGeometry(QtCore.QRect(390, 460, 141, 23))
        self.UnfollowButton.setStyleSheet("border-radius:5px;\n"
"background-color: rgb(79, 74, 84);\n"
"color: rgb(255, 255, 255);")
        self.UnfollowButton.setObjectName("UnfollowButton")
        self.label_31 = QtWidgets.QLabel(self.widget)
        self.label_31.setGeometry(QtCore.QRect(305, 730, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.label_31.setFont(font)
        self.label_31.setStyleSheet("color: rgb(255, 255, 255);")
        self.label_31.setObjectName("label_31")
        self.Unfollowed = QtWidgets.QLabel(self.widget)
        self.Unfollowed.setGeometry(QtCore.QRect(495, 730, 41, 20))
        font = QtGui.QFont()
        font.setPointSize(11)
        self.Unfollowed.setFont(font)
        self.Unfollowed.setStyleSheet("color: rgb(255, 255, 255);")
        self.Unfollowed.setAlignment(QtCore.Qt.AlignCenter)
        self.Unfollowed.setObjectName("Unfollowed")
        self.BG.raise_()
        self.label.raise_()
        self.LoginIGmainLabel.raise_()
        self.label_2.raise_()
        self.label_3.raise_()
        self.loginEntry.raise_()
        self.passwordEntry.raise_()
        self.saveButton.raise_()
        self.loadButton.raise_()
        self.delButton.raise_()
        self.label_4.raise_()
        self.label_5.raise_()
        self.FirefoxButton.raise_()
        self.Chromebutton.raise_()
        self.label_6.raise_()
        self.LoginIGmainLabel_2.raise_()
        self.IloscKontEntry.raise_()
        self.NazwaKontaEntry.raise_()
        self.label_7.raise_()
        self.label_8.raise_()
        self.FollowersButton.raise_()
        self.label_11.raise_()
        self.NazwaHaszEntry.raise_()
        self.HaszButton.raise_()
        self.LoginIGmainLabel_3.raise_()
        self.label_12.raise_()
        self.label_13.raise_()
        self.IloscHaszKont.raise_()
        self.label_15.raise_()
        self.LoginIGmainLabel_4.raise_()
        self.ComentCheckBox.raise_()
        self.LikeCheckBox.raise_()
        self.label_16.raise_()
        self.label_17.raise_()
        self.FollowButton.raise_()
        self.label_18.raise_()
        self.LoginIGmainLabel_5.raise_()
        self.Message_Button.raise_()
        self.label_9.raise_()
        self.LoginIGmainLabel_6.raise_()
        self.label_19.raise_()
        self.label_10.raise_()
        self.label_20.raise_()
        self.CommentEdit.raise_()
        self.MessageEdit.raise_()
        self.Website.raise_()
        self.label_21.raise_()
        self.label_22.raise_()
        self.timerLabel.raise_()
        self.label_24.raise_()
        self.followCountLabel.raise_()
        self.label_26.raise_()
        self.DownloadedLabel.raise_()
        self.label_27.raise_()
        self.SendedMessage.raise_()
        self.label_25.raise_()
        self.label_28.raise_()
        self.LikedPhotoLabel.raise_()
        self.ComentLabel.raise_()
        self.CloseButton.raise_()
        self.minimaliseButton.raise_()
        self.label_14.raise_()
        self.label_29.raise_()
        self.WhoIsLogin.raise_()
        self.label_30.raise_()
        self.HowMuchDays.raise_()
        self.Ver.raise_()
        self.LoginIGmainLabel_7.raise_()
        self.DodajKonto.raise_()
        self.UnfollowButton.raise_()
        self.label_31.raise_()
        self.Unfollowed.raise_()
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        MainWindow.setTabOrder(self.loginEntry, self.passwordEntry)
        MainWindow.setTabOrder(self.passwordEntry, self.saveButton)
        MainWindow.setTabOrder(self.saveButton, self.loadButton)
        MainWindow.setTabOrder(self.loadButton, self.delButton)
        MainWindow.setTabOrder(self.delButton, self.FirefoxButton)
        MainWindow.setTabOrder(self.FirefoxButton, self.Chromebutton)
        MainWindow.setTabOrder(self.Chromebutton, self.NazwaKontaEntry)
        MainWindow.setTabOrder(self.NazwaKontaEntry, self.IloscKontEntry)
        MainWindow.setTabOrder(self.IloscKontEntry, self.FollowersButton)
        MainWindow.setTabOrder(self.FollowersButton, self.NazwaHaszEntry)
        MainWindow.setTabOrder(self.NazwaHaszEntry, self.IloscHaszKont)
        MainWindow.setTabOrder(self.IloscHaszKont, self.HaszButton)
        MainWindow.setTabOrder(self.HaszButton, self.ComentCheckBox)
        MainWindow.setTabOrder(self.ComentCheckBox, self.LikeCheckBox)
        MainWindow.setTabOrder(self.LikeCheckBox, self.FollowButton)
        MainWindow.setTabOrder(self.FollowButton, self.Message_Button)
        MainWindow.setTabOrder(self.Message_Button, self.DodajKonto)
        MainWindow.setTabOrder(self.DodajKonto, self.UnfollowButton)
        MainWindow.setTabOrder(self.UnfollowButton, self.Website)
        MainWindow.setTabOrder(self.Website, self.MessageEdit)
        MainWindow.setTabOrder(self.MessageEdit, self.CommentEdit)
        MainWindow.setTabOrder(self.CommentEdit, self.minimaliseButton)
        MainWindow.setTabOrder(self.minimaliseButton, self.CloseButton)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.LoginIGmainLabel.setText(_translate("MainWindow", "Dane do konta IG"))
        self.label_2.setText(_translate("MainWindow", "Login:"))
        self.label_3.setText(_translate("MainWindow", "Hasło:"))
        self.saveButton.setText(_translate("MainWindow", "Zapisz Konto"))
        self.loadButton.setText(_translate("MainWindow", "Załaduj Konto"))
        self.delButton.setText(_translate("MainWindow", "Usuń Konto"))
        self.label_4.setText(_translate("MainWindow", "Firefox"))
        self.label_5.setText(_translate("MainWindow", "Chrome"))
        self.LoginIGmainLabel_2.setText(_translate("MainWindow", "Pobierz obserwujących profil"))
        self.label_7.setText(_translate("MainWindow", "Nazwa Konta:"))
        self.label_8.setText(_translate("MainWindow", "Ilość kont:"))
        self.FollowersButton.setText(_translate("MainWindow", "Uruchom Pobieranie"))
        self.HaszButton.setText(_translate("MainWindow", "Uruchom Pobieranie"))
        self.LoginIGmainLabel_3.setText(_translate("MainWindow", "Pobierz konta #"))
        self.label_12.setText(_translate("MainWindow", "#"))
        self.label_13.setText(_translate("MainWindow", "Ilość kont:"))
        self.LoginIGmainLabel_4.setText(_translate("MainWindow", "Obserwacja z Pliku"))
        self.label_16.setText(_translate("MainWindow", "Dodaj komentarz"))
        self.label_17.setText(_translate("MainWindow", "Polub zdjęcie"))
        self.FollowButton.setText(_translate("MainWindow", "Uruchom Obserowanie"))
        self.LoginIGmainLabel_5.setText(_translate("MainWindow", "Wyślij Wiadomość"))
        self.Message_Button.setText(_translate("MainWindow", "Uruchom Rozmowe"))
        self.LoginIGmainLabel_6.setText(_translate("MainWindow", "INFORMACJE"))
        self.label_19.setText(_translate("MainWindow", "Status:"))
        self.label_20.setText(_translate("MainWindow", "Nieaktywny"))
        self.CommentEdit.setText(_translate("MainWindow", "Edytuj Komentarze"))
        self.MessageEdit.setText(_translate("MainWindow", "Edytuj Wiadomości"))
        self.Website.setText(_translate("MainWindow", "Kontakt z Nami"))
        self.label_21.setText(_translate("MainWindow", "© Daniel Zieliński, Artur Kaczuba, Łukasz Mela 2021."))
        self.label_22.setText(_translate("MainWindow", "Do następnej sesji zostało:"))
        self.timerLabel.setText(_translate("MainWindow", "0s"))
        self.label_24.setText(_translate("MainWindow", "Zaobserwowanych:"))
        self.followCountLabel.setText(_translate("MainWindow", "0"))
        self.label_26.setText(_translate("MainWindow", "Pobranych Nazw Kont:"))
        self.DownloadedLabel.setText(_translate("MainWindow", "0"))
        self.label_27.setText(_translate("MainWindow", "Wysłano wiadomości:"))
        self.SendedMessage.setText(_translate("MainWindow", "0"))
        self.label_25.setText(_translate("MainWindow", "Dodanych Komentarzy:"))
        self.label_28.setText(_translate("MainWindow", "Polubionych Zdjęć:"))
        self.LikedPhotoLabel.setText(_translate("MainWindow", "0"))
        self.ComentLabel.setText(_translate("MainWindow", "0"))
        self.label_14.setText(_translate("MainWindow", "Ver:"))
        self.label_29.setText(_translate("MainWindow", "Zalogowany:"))
        self.WhoIsLogin.setText(_translate("MainWindow", "TESTER"))
        self.label_30.setText(_translate("MainWindow", "Pozostało:"))
        self.HowMuchDays.setText(_translate("MainWindow", "999 Dni"))
        self.Ver.setText(_translate("MainWindow", "testowa"))
        self.LoginIGmainLabel_7.setText(_translate("MainWindow", "Unfollow"))
        self.DodajKonto.setText(_translate("MainWindow", "Zapisz Konta"))
        self.UnfollowButton.setText(_translate("MainWindow", "Uruchom Unfollow"))
        self.label_31.setText(_translate("MainWindow", "Odobserowanych:"))
        self.Unfollowed.setText(_translate("MainWindow", "0"))


class Popup(object):
    def setup(self, Form):
        Form.setObjectName("Form")
        Form.resize(738, 278)
        Form.setMinimumSize(QtCore.QSize(738, 278))
        Form.setMaximumSize(QtCore.QSize(738, 278))
        Form.setStyleSheet("")
        Form.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        Form.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(50, 30, 600, 201))
        self.label.setStyleSheet("background-color: rgb(40,36,46);\n"
                                 "border-radius:50px;\n"
                                 "")
        self.label.setText("")
        self.label.setObjectName("label")
        self.label_10 = QtWidgets.QLabel(Form)
        self.label_10.setGeometry(QtCore.QRect(280, 30, 141, 51))
        self.label_10.setStyleSheet("border-image: url(:/logo/logo white bez tla.png);")
        self.label_10.setText("")
        self.label_10.setObjectName("label_10")
        self.confirmButton = QtWidgets.QToolButton(Form)
        self.confirmButton.setGeometry(QtCore.QRect(310, 180, 81, 23))
        self.confirmButton.setStyleSheet("border-radius:5px;\n"
                                         "background-color: rgb(79, 74, 84);\n"
                                         "color: rgb(255, 255, 255);")
        self.confirmButton.setObjectName("confirmButton")
        self.label_1 = QtWidgets.QLabel(Form)
        self.label_1.setGeometry(QtCore.QRect(80, 80, 541, 91))
        self.label_1.setStyleSheet("background-color: rgb(53,48,61);\n"
                                   "color:rgb(255, 255, 255);\n"
                                   "border-radius:20px;")
        self.label_1.setText("")
        self.label_1.setAlignment(QtCore.Qt.AlignCenter)
        self.label_1.setObjectName("label_1")
        self.popupTitle = QtWidgets.QLabel(Form)
        self.popupTitle.setGeometry(QtCore.QRect(80, 90, 541, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.popupTitle.setFont(font)
        self.popupTitle.setStyleSheet("color:rgb(255, 255, 255);")
        self.popupTitle.setAlignment(QtCore.Qt.AlignCenter)
        self.popupTitle.setObjectName("popupTitle")
        self.popupContent = QtWidgets.QLabel(Form)
        self.popupContent.setGeometry(QtCore.QRect(80, 119, 541, 41))
        self.popupContent.setStyleSheet("color:rgb(255,255,255);")
        self.popupContent.setAlignment(QtCore.Qt.AlignCenter)
        self.popupContent.setObjectName("popupContent")

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.confirmButton.setText(_translate("Form", "Ok"))
        self.popupTitle.setText(_translate("Form", "popupTitle"))
        self.popupContent.setText(_translate("Form", "popupContent"))

        # self.confirmButton.clicked.connect(self.clos)


class decPopup(object):
    def decSetup(self, Form):
        Form.setObjectName("Form")
        Form.resize(738, 278)
        Form.setMinimumSize(QtCore.QSize(738, 278))
        Form.setMaximumSize(QtCore.QSize(738, 278))
        Form.setStyleSheet("")
        Form.setWindowFlags(QtCore.Qt.FramelessWindowHint)
        Form.setAttribute(QtCore.Qt.WA_TranslucentBackground)
        self.label = QtWidgets.QLabel(Form)
        self.label.setGeometry(QtCore.QRect(50, 30, 600, 201))
        self.label.setStyleSheet("background-color: rgb(40,36,46);\n"
                                 "border-radius:50px;\n"
                                 "")
        self.label.setText("")
        self.label.setObjectName("label")
        self.label_10 = QtWidgets.QLabel(Form)
        self.label_10.setGeometry(QtCore.QRect(280, 30, 141, 51))
        self.label_10.setStyleSheet("border-image: url(:/logo/logo white bez tla.png);")
        self.label_10.setText("")
        self.label_10.setObjectName("label_10")
        self.yesButton = QtWidgets.QToolButton(Form)
        self.yesButton.setGeometry(QtCore.QRect(265, 180, 81, 23))
        self.yesButton.setStyleSheet("border-radius:5px;\n"
                                     "background-color: rgb(79, 74, 84);\n"
                                     "color: rgb(255, 255, 255);")
        self.yesButton.setObjectName("yesButton")
        self.label_1 = QtWidgets.QLabel(Form)
        self.label_1.setGeometry(QtCore.QRect(80, 80, 541, 91))
        self.label_1.setStyleSheet("background-color: rgb(53,48,61);\n"
                                   "color:rgb(255, 255, 255);\n"
                                   "border-radius:20px;")
        self.label_1.setText("")
        self.label_1.setAlignment(QtCore.Qt.AlignCenter)
        self.label_1.setObjectName("label_1")
        self.dPopupTitle = QtWidgets.QLabel(Form)
        self.dPopupTitle.setGeometry(QtCore.QRect(80, 90, 541, 16))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.dPopupTitle.setFont(font)
        self.dPopupTitle.setStyleSheet("color:rgb(255, 255, 255);")
        self.dPopupTitle.setAlignment(QtCore.Qt.AlignCenter)
        self.dPopupTitle.setObjectName("dPopupTitle")
        self.dPopupContent = QtWidgets.QLabel(Form)
        self.dPopupContent.setGeometry(QtCore.QRect(80, 119, 541, 41))
        self.dPopupContent.setStyleSheet("color:rgb(255,255,255);")
        self.dPopupContent.setAlignment(QtCore.Qt.AlignCenter)
        self.dPopupContent.setObjectName("dPopupContent")
        self.noButton = QtWidgets.QToolButton(Form)
        self.noButton.setGeometry(QtCore.QRect(355, 180, 81, 23))
        self.noButton.setStyleSheet("border-radius:5px;\n"
                                    "background-color: rgb(79, 74, 84);\n"
                                    "color: rgb(255, 255, 255);")
        self.noButton.setObjectName("noButton")

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.yesButton.setText(_translate("Form", "Tak"))
        self.dPopupTitle.setText(_translate("Form", "popupTitle"))
        self.dPopupContent.setText(_translate("Form", "popupContent"))
        self.noButton.setText(_translate("Form", "Nie"))


class ComEdit(object):

        def ComEditsetup(self, Form):
                Form.setObjectName("Form")
                Form.resize(532, 657)
                Form.setMinimumSize(QtCore.QSize(532, 657))
                Form.setMaximumSize(QtCore.QSize(532, 657))
                Form.setStyleSheet("")
                Form.setWindowFlags(QtCore.Qt.FramelessWindowHint)
                Form.setAttribute(QtCore.Qt.WA_TranslucentBackground)
                self.label = QtWidgets.QLabel(Form)
                self.label.setGeometry(QtCore.QRect(20, 20, 420, 561))
                self.label.setStyleSheet("background-color: rgb(40,36,46);\n"
                                         "border-radius:50px;\n"
                                         "")
                self.label.setText("")
                self.label.setObjectName("label")
                self.CloseButton = QtWidgets.QToolButton(Form)
                self.CloseButton.setGeometry(QtCore.QRect(380, 40, 16, 16))
                font = QtGui.QFont()
                font.setFamily("Montserrat")
                font.setPointSize(7)
                font.setBold(False)
                font.setWeight(50)
                self.CloseButton.setFont(font)
                self.CloseButton.setToolTipDuration(1)
                self.CloseButton.setLayoutDirection(QtCore.Qt.LeftToRight)
                self.CloseButton.setStyleSheet("background-color: rgb(255, 96, 92);\n"
                                               "image: url(:/ex/close.png);\n"
                                               "border-radius:8px;\n"
                                               "")
                self.CloseButton.setText("")
                self.CloseButton.setIconSize(QtCore.QSize(25, 25))
                self.CloseButton.setToolButtonStyle(QtCore.Qt.ToolButtonIconOnly)
                self.CloseButton.setAutoRaise(False)
                self.CloseButton.setObjectName("CloseButton")
                self.minimaliseButton = QtWidgets.QToolButton(Form)
                self.minimaliseButton.setGeometry(QtCore.QRect(360, 40, 16, 16))
                self.minimaliseButton.setStyleSheet("background-color: rgb(255, 189, 68);\n"
                                                    "image: url(:/min/minimize.png);\n"
                                                    "border-radius:8px;\n"
                                                    "")
                self.minimaliseButton.setText("")
                self.minimaliseButton.setIconSize(QtCore.QSize(25, 25))
                self.minimaliseButton.setObjectName("minimaliseButton")
                self.label_10 = QtWidgets.QLabel(Form)
                self.label_10.setGeometry(QtCore.QRect(160, 30, 141, 51))
                self.label_10.setStyleSheet("border-image: url(:/logo/logo white bez tla.png);")
                self.label_10.setText("")
                self.label_10.setObjectName("label_10")
                self.Website = QtWidgets.QToolButton(Form)
                self.Website.setGeometry(QtCore.QRect(160, 530, 141, 23))
                self.Website.setStyleSheet("border-radius:5px;\n"
                                           "background-color: rgb(79, 74, 84);\n"
                                           "color: rgb(255, 255, 255);")
                self.Website.setObjectName("Website")
                self.LoginIGmainLabel = QtWidgets.QLabel(Form)
                self.LoginIGmainLabel.setGeometry(QtCore.QRect(160, 90, 141, 16))
                font = QtGui.QFont()
                font.setPointSize(11)
                self.LoginIGmainLabel.setFont(font)
                self.LoginIGmainLabel.setStyleSheet("color: rgb(255, 255, 255);")
                self.LoginIGmainLabel.setAlignment(QtCore.Qt.AlignCenter)
                self.LoginIGmainLabel.setObjectName("LoginIGmainLabel")
                self.textEdit = QtWidgets.QTextEdit(Form)
                self.textEdit.setGeometry(QtCore.QRect(43, 120, 371, 391))
                self.textEdit.setStyleSheet("background-color: rgb(79,74,84);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border: none;\n"
                                            "border-radius:10px;\n"
                                            "\n"
                                            "")
                self.textEdit.setObjectName("textEdit")

                self.retranslateUi(Form)
                QtCore.QMetaObject.connectSlotsByName(Form)

        def retranslateUi(self, Form):
                _translate = QtCore.QCoreApplication.translate
                Form.setWindowTitle(_translate("Form", "Form"))
                self.Website.setText(_translate("Form", "Zapisz"))
                self.LoginIGmainLabel.setText(_translate("Form", "Zmień Komentarze"))


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent=parent)
        ui = Ui_MainWindow()
        ui.setupUi(self)
        ## BUTTONS ##
        ui.CloseButton.clicked.connect(self.close)
        ui.minimaliseButton.clicked.connect(self.minimized)
        ui.saveButton.clicked.connect(self.save)
        ui.loadButton.clicked.connect(self.load)
        ui.delButton.clicked.connect(self.delete)
        ui.FollowersButton.clicked.connect(self.StartFollowedDownload)
        ui.HaszButton.clicked.connect(self.StartHaszDownload)
        ui.FollowButton.clicked.connect(self.StartFollowMaker)
        ui.Message_Button.clicked.connect(self.StartMessageSend)
        ui.Website.clicked.connect(self.ContactUS)
        ui.MessageEdit.clicked.connect(self.mesEdit)
        ui.CommentEdit.clicked.connect(self.comEdit)
        ui.DodajKonto.clicked.connect(self.addAccount)
        ui.UnfollowButton.clicked.connect(self.StartUnfollow)

        global l, p, accname, accCount, haszname, haszCount, coment, liked, lista, clock, followed_count, liked_count, comment_count, downloaded, sendmsg, F, C, status, \
            popupTitle, PopupContent, dPopupTitle, dPopupContent, x, x1, unf
        ## ENTRY GET ##
        l = ui.loginEntry
        p = ui.passwordEntry

        ## Pobierz obserwowanych
        accname = ui.NazwaKontaEntry
        accCount = ui.IloscKontEntry

        ## Pobierz hasz
        haszname = ui.NazwaHaszEntry
        haszCount = ui.IloscHaszKont

        ## Follow
        coment = ui.ComentCheckBox
        liked = ui.LikeCheckBox

        ##Message
        ##BRAK

        ## Info

        clock = ui.timerLabel
        followed_count = ui.followCountLabel
        liked_count = ui.LikedPhotoLabel
        comment_count = ui.ComentLabel
        downloaded = ui.DownloadedLabel
        sendmsg = ui.SendedMessage
        status = ui.label_20
        unf = ui.Unfollowed

        ##WebBrowser
        F = ui.FirefoxButton
        C = ui.Chromebutton


        ## Wersja/Who/data
        ui.Ver.setText(ver)
        ui.WhoIsLogin.setText("BetaTester")
        ui.HowMuchDays.setText("14 Dni")

    ### Funkcjonalności pozostałych Przycisków

    def close(self):
        app.exit()

    def minimized(self):
        self.showMinimized()

    def mousePressEvent(self, event):  # +
        self.dragPos = event.globalPos()

    def mouseMoveEvent(self, event):  # !!!
        if event.buttons() == QtCore.Qt.LeftButton:
            self.move(self.pos() + event.globalPos() - self.dragPos)
            self.dragPos = event.globalPos()
            event.accept()

    def stopBot(self):
        print("ne")

    ### Zapisz/Załaduj/Usuń Dane o koncie IG
    def save(self):
        global dPopupTitle, dPopupContent, popupTitle, popupContent
        login = l.text()
        password = p.text()
        who = "save"
        if len(login) == 0 or len(password) == 0:
            popupTitle = "Zapisz Konto"
            popupContent = "Uzupełnij dane o koncie IG"
            self.popup(popupTitle, popupContent)


        elif os.path.exists("Res/user.daz"):
            dPopupTitle = "Zapisz Konto"
            dPopupContent = "Czy chcesz nadpisać dane o koncie Instagram?"
            self.decidePopup(dPopupTitle, dPopupContent, who)


        else:
            f = open("Res/user.daz", "w")
            f.write(login + "\n" + password)
            f.close()
            print("L: " + login + " P: " + password)
            popupTitle = "Zapisz Konto"
            popupContent = "Konto zostało zapisane"
            self.popup(popupTitle, popupContent)

    def YesSave(self):
        global dPopupTitle, dPopupContent, popupTitle, popupContent
        login = l.text()
        password = p.text()
        f = open("Res/user.daz", "w")
        f.write(login + "\n" + password)
        f.close()
        print("L: " + login + " P: " + password)
        popupTitle = "Zapisz Konto"
        popupContent = "Konto zostało nadpisane"
        self.popup(popupTitle, popupContent)

    def NoSave(self):
        popupTitle = "Zapisz Konto"
        popupContent = "Konto nie zostało nadpisane"
        self.popup(popupTitle, popupContent)

    def load(self):
        global dPopupTitle, dPopupContent, popupTitle, popupContent
        if os.path.exists("Res/user.daz"):
            f = open("Res/user.daz", "r")
            lo = str(f.readlines(1))
            LenL = len(lo) - 4
            login = lo[2:LenL]
            if len(login) == 0:
                popupTitle = "Załaduj Konto"
                popupContent = "Brak loginu w pliku user.daz"
                self.popup(popupTitle, popupContent)
            else:
                l.clear()
                l.setText(login)

                pw = str(f.readlines(2))
                LenP = len(pw) - 2
                passwd = pw[2:LenP]
                if len(passwd) == 0:
                    popupTitle = "Załaduj Konto"
                    popupContent = "Brak hasła w pliku user.daz"
                    self.popup(popupTitle, popupContent)
                else:
                    p.clear()
                    p.setText(passwd)
        else:
            popupTitle = "Załaduj Konto"
            popupContent = "Brak pliku z danymi do logowania"
            self.popup(popupTitle, popupContent)

    def delete(self):
        global dPopupTitle, dPopupContent, popupTitle, popupContent, x
        if os.path.exists("Res/user.daz"):
            who = "del"
            dPopupTitle = "Usuń Konto"
            dPopupContent = "Czy chcesz usunąć dane o koncie Instagram?"
            self.decidePopup(dPopupTitle, dPopupContent, who)

        else:

            popupTitle = "Usuń Konto"
            popupContent = "Brak pliku z zapisanym kontem"
            self.popup(popupTitle, popupContent)

    def YesDel(self):
        os.remove("Res/user.daz")
        l.clear()
        p.clear()

        popupTitle = "Usuń Konto"
        popupContent = "Plik user.daz został pomyślnie usunięty"
        self.popup(popupTitle, popupContent)

    def NoDel(self):
        popupTitle = "Usuń Konto"
        popupContent = "Plik nie został usunięty"
        self.popup(popupTitle, popupContent)

    ### Thread do funkcji

    def thread_FollowedDownload(self):
        global t1
        t1 = Thread(target=self.FollowedDownload)
        t1.start()

    def thread_HaszDownload(self):
        global t2
        t2 = Thread(target=self.HaszDownload)
        t2.start()

    def thread_Follow(self):
        global t3
        t3 = Thread(target=self.FollowMaker)
        t3.start()

    def thread_Message(self):
        global t4
        t4 = Thread(target=self.MessageSend)
        t4.start()

    def thread_Unfollow(self):
        global t5
        t5 = Thread(target=self.Unfollow)
        t5.start()

    ### Pobierz obserwujących

    def StartFollowedDownload(self):
        global problem_text
        ### Ustawienie Aktywny w Info
        status.setText("Aktywny")
        status.setStyleSheet('color: rgb(102, 137, 37) ')
        QApplication.processEvents()
        clock.setText("0 s")
        ############DANE##############
        profiles = accname.text()
        howmuchprofil = accCount.text()
        log = l.text()
        pas = p.text()
        d = "qwertyuiop[]asdfghjl;'zxcvbnm,./{}:<>?" + '"'
        print(howmuchprofil[0:1])
        if len(log) == 0 or len(pas) == 0:
            popupTitle = "Insta Login"
            popupContent = "Uzupełni Dane do konta IG"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

        if len(howmuchprofil) == 0 or len(profiles) == 0:
            popupTitle = "Pobierz Obserwujących Profil"
            popupContent = "Uzupełnij pola Nazwa Konta i Ilość Kont"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

        elif (howmuchprofil[0:1] in d) == True:
            popupTitle = "Pobierz Obserwujących Profil"
            popupContent = "W polu Ilość Kont nie ma wartości liczbowej"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

        else:
            ###SPRAWDZAM WYPELNIONE POLA###
            self.thread_FollowedDownload()

    def FollowedDownload(self):
        global problem_text
        profile = accname.text()
        howmuchprofil = accCount.text()
        ####Wartości Początkowe####
        problem = 0
        hMPGet = int(howmuchprofil)
        howMuchProfile = hMPGet
        maxloop = round(howMuchProfile / 10) + 2
        name = ""

        #####Skrypt#####

        self.instaLogin()
        bot.get('https://www.instagram.com/' + profile)
        sleep(randint(2, 6))
        try:
            bot.find_element_by_xpath("//*[text()='Przepraszamy, ta strona jest niedostępna']")
            problem = 1
            self.problem_Status()
            bot.close()
            sys.exit(1)
        except:
            try:
                bot.find_element_by_xpath("//*[text()=' obserwujących']").click()
            except:
                problem = 1
                problem_text = "Problem ze znalezieniem przycisku listy obserwujących #FollowedDownload"
                self.problem()
                bot.close()
                sys.exit(1)

            sleep(randint(2, 6))
        if os.path.exists("DataProfile.xlsx"):
            workbook = openpyxl.load_workbook('DataProfile.xlsx')
            worksheet = workbook.active
            worksheet.delete_cols(1, 5)

        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "DATA"
            workbook.save(filename='DataProfile.xlsx')

        worksheet.cell(row=1, column=1, value="LP.")
        worksheet.cell(row=1, column=2, value="Nazwa Konta")
        worksheet.cell(row=1, column=3, value="Link do profilu")
        sleep(randint(2, 6))
        fBody = bot.find_element_by_xpath("//div[@class='isgrP']")
        profmin = 1
        profmax = 10

        for i in range(1, maxloop):

            for j in range(0, 3):
                bot.execute_script(
                    'arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].offsetHeight;',
                    fBody)
                sleep(randint(2, 6))

            if (i == 1):
                profMin = profmin
                profMax = profmax
            else:
                profMax = profMax + profmax
            for k in range(profMin, profMax):
                if (k == howMuchProfile + 1):
                    break
                try:
                    namesearch = bot.find_element_by_xpath(
                        "/html/body/div[5]/div/div/div[2]/ul/div/li[" + str(k) + "]/div/div[1]/div[2]/div[1]/span/a")
                    name = namesearch.text
                except:
                    print("Spróbuje jeszcze raz")
                    try:
                        namess = bot.find_element_by_xpath('/html/body/div[5]/div/div/div[2]/ul/div/li[' + str(
                            k) + ']/div/div[1]/div[2]/div[1]/span/a')
                        name = namess.text
                    except:
                        try:
                            namess = bot.find_element_by_xpath('/html/body/div[6]/div/div/div[2]/ul/div/li[' + str(
                                k) + ']/div/div[1]/div[2]/div[1]/span/a')
                            name = namess.text
                        except:
                            problem = 1
                            problem_text = "Nie mogłem pobrać informacji o koncie #FollowedDownload"
                            self.problem()
                num = k
                print(name)
                print(num)
                downloaded.setText(str(num))
                worksheet.cell(row=k + 1, column=1, value=num)
                worksheet.cell(row=k + 1, column=2, value=name)
                worksheet.cell(row=k + 1, column=3, value="https://www.instagram.com/" + name)
            profMin = profMin + profmax - 1
            workbook.save('DataProfile.xlsx')
            sleep(randint(2, 6))

        ### Ustawienie Zakończono w Info
        if problem == 0:
            bot.close()
            QApplication.processEvents()
            status.setStyleSheet('color: rgb(227, 62, 51);')
            status.setText("Zakończono")

    ### Pobierz używających Hasz

    def StartHaszDownload(self):
        ### Ustawienie Aktywny w Info
        global problem_text
        ### Ustawienie Aktywny w Info
        status.setText("Aktywny")
        status.setStyleSheet('color: rgb(102, 137, 37) ')
        clock.setText("0 s")
        QApplication.processEvents()

        ############DANE##############
        hasztag = haszname.text()
        hasznum = haszCount.text()
        log = l.text()
        pas = p.text()
        d = "qwertyuiop[]asdfghjl;'zxcvbnm,./{}:<>?" + '"'
        ##SPRAWDZANIE
        if len(log) == 0 or len(pas) == 0:
            popupTitle = "Insta Login"
            popupContent = "Uzupełni Dane do konta IG"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

        if len(hasztag) == 0 or len(hasznum) == 0:
            popupTitle = "Pobierz Użytkowników Hasztag"
            popupContent = "Uzupełnij pola Nazwa # i Ilość kont"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

        elif (hasznum[0:1] in d) == True:
            popupTitle = "Pobierz Obserwujących Profil"
            popupContent = "Pole"
            self.popup(popupTitle, popupContent)
            self.problem_Status()
        else:
            self.thread_HaszDownload()

    def HaszDownload(self):
        global problem_text

        ############DANE##############
        hasztag = haszname.text()
        hasznum = int(haszCount.text())
        hasznumb = int((hasznum / 200) + 1)

        ###############Wartości Początkowe######
        maxhasz = 200
        problem = 0
        h = 0
        numberPost = 0
        href = ""
        ### SCRYPTY ###
        profiler = '//*[@class="g47SY "]'
        ### Sprawdzanie i Tworzenie raportu
        if os.path.exists('HaszProfile.xlsx'):
            workbook = openpyxl.load_workbook('HaszProfile.xlsx')
            worksheet = workbook.active
            worksheet.delete_cols(1, 5)

        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "#Pofiles"
            workbook.save(filename='HaszProfile.xlsx')

        worksheet.cell(row=1, column=1, value="LP.")
        worksheet.cell(row=1, column=2, value="Nazwa Konta")
        worksheet.cell(row=1, column=3, value="Link do profilu")

        self.instaLogin()
        bot.get("https://www.instagram.com/explore/tags/" + hasztag + "/")
        sleep(4)

        try:
            numberPost = bot.find_element_by_xpath(profiler).text

        except:
            numberPost = hasznumb
        try:
            bot.find_element_by_xpath('/html/body/div[1]/section/main/article/div[1]/div/div/div[1]/div[1]').click()
        except:
            problem = 1
            problem_text = "Problem ze zlokalizowaniem i naciśnięciem 1 zdjęcia#HaszDownload"
            self.problem()

        sleep(randint(3, 6))
        hmax = hasznum - 1
        print("pobieranie")
        print(numberPost)
        for i in range(0, hasznumb):
            if i == 0:
                minhasz = 0
                maxhaszb = maxhasz
                print("i0")
            else:
                if (j == numberPost):
                    break
                else:
                    print("i1+")
                    maxhaszb = maxhaszb + maxhasz
                    timeslow = (60 * 10) - 1
                    for remaining in range(timeslow, 0, -1):
                        clock.setText(str(remaining) + " s")
                        sleep(1)
                    clock.setText("0 s")

            for j in range(minhasz, maxhaszb):
                if j == hmax:
                    break
                elif j >= hmax:
                    break
                elif j == numberPost:
                    break
                else:
                    try:
                        namesearch = bot.find_element_by_xpath(
                            "/html/body/div[5]/div[2]/div/article/header/div[2]/div[1]/div/span/a")
                        href = namesearch.get_attribute('href')
                        h += 1


                    except:

                        try:
                            namesearch = bot.find_element_by_xpath(
                                '/html/body/div[6]/div[2]/div/article/header/div[2]/div[1]/div/span/a')
                            href = namesearch.get_attribute('href')
                            h += 1
                        except:
                            try:
                                namesearch = bot.find_element_by_xpath(
                                    '/html/body/div[6]/div[2]/div/article/header/div[2]/div[1]/div/span/a')
                                href = namesearch.text
                                h += 1

                            except:

                                print("problem z pobraniem nazwy")

                    num = str(j + 1)
                    if len(href) == 0:
                        break
                    else:
                        lenhf = len(href) - 1
                        name = href[26:lenhf]
                        worksheet.cell(row=j + 2, column=1, value=num)
                        worksheet.cell(row=j + 2, column=2, value=name)
                        worksheet.cell(row=j + 2, column=3, value=href)
                    try:
                        bot.find_element_by_css_selector('._65Bje').click()
                    except:
                        try:
                            bot.find_element_by_xpath("//*[text()='Dalej']")
                        except:
                            problem_text = "Problem ze zlokalizowaniem przycisku dalej #HaszDownload"
                            self.perror()
                    sleep(1)
                    print(h)
                    downloaded.setText(str(h))

            minhasz = maxhaszb
            workbook.save(filename='HaszProfile.xlsx')
        dt = pd.read_excel('HaszProfile.xlsx')
        a = dt.drop_duplicates(subset=['Nazwa Konta', 'Link do profilu'], keep='first')
        df = pd.DataFrame(a, columns=['Nazwa Konta', 'Link do profilu'])
        df.to_excel('HaszProfile.xlsx', sheet_name='#Profiles')

        ### Ustawienie Zakończono w Info
        if problem == 0:
            bot.close()
            QApplication.processEvents()
            status.setStyleSheet('color: rgb(227, 62, 51);')
            status.setText("Zakończono")

    ### Zaobserwuj Ludzi polub zdjęcie i dodaj Komentarz
    def StartFollowMaker(self):
        ### Ustawienie Aktywny w Info
        status.setText("Aktywny")
        status.setStyleSheet('color: rgb(102, 137, 37) ')
        QApplication.processEvents()

        log = l.text()
        pas = p.text()

        ##SPRAWDZANIE
        if len(log) == 0 or len(pas) == 0:
            popupTitle = "Insta Login"
            popupContent = "Uzupełni Dane do konta IG"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

        if os.path.exists("DataProfile.xlsx"):
            self.thread_Follow()

        else:
            popupTitle = "Obserwacja z Pliku"
            popupContent = "Utwórz plik o nazwie DataProfile.xlsx lub pobierz obserwujące profile"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

    def FollowMaker(self):
        ### wartości globalne informacji
        global problem_text
        ### Ustawienie Aktywny w Info
        status.setText("Aktywny")
        status.setStyleSheet('color: rgb(102, 137, 37) ')
        QApplication.processEvents()

        ############SKRYPCIKI#############
        obserwuj = "//*[text()='Obserwuj']"
        wyslaneZaproszenie = "//*[text()='Wysłane zaproszenie']"
        wyslijWiadomosc = "//*[text()='Wyślij wiadomość']"
        rowniezObserwuj = "//*[text()='Również Obserwuj']"
        like = '/html/body/div[5]/div[2]/div/article/div[3]/section[1]/span[1]/button'  # '//*[@aria-label="Lubię to!"]'
        nextphoto = '//*[text()="Dalej"]'
        koment = '//*[@aria-label="Dodaj komentarz..."]'
        publish = '//*[@type="submit"]'
        ###############Wartości Początkowe######
        slowwly = 0
        follow = 0
        lc = 0  # LIKE COUNT
        cm = 0  # COMENT COUNT
        minlc = 2
        mpc = 6
        ## USTAWIANIE Wartości
        com = coment.isChecked()
        likeed = liked.isChecked()
        timer = clock

        # print(com.isChecked(), likeed.isChecked())
        #### KOMENTARZE Z PLIKU ####
        file = open("Res/Komentarz.daz", encoding="utf-8")
        lines = file.read()

        com_list = lines.splitlines()
        file.close()
        com_com = len(com_list)
        ##LOGOWANIE##
        self.instaLogin()

        if os.path.exists('DataProfile.xlsx'):
            workbook = openpyxl.load_workbook('DataProfile.xlsx')
            worksheet = workbook.active
            worksheet.cell(row=1, column=4, value='Informacja o Obserwacji')
            worksheet.cell(row=1, column=5, value='Informacja o Koncie')
            worksheet.cell(row=1, column=6, value='Polubiono zdjęć')
            worksheet.cell(row=1, column=7, value='Komentarz do zdjęcia')
            worksheet.cell(row=1, column=8, value='Umieszczony komentarz')

            # Ile razy wykona sie pętla główna
            Loopmax = 14
            for i in range(1, Loopmax):
                print("Pętla " + str(i) + " z " + str(Loopmax))
                maxF30 = randint(17, 21)
                print("Zostanie zaobserwowanych " + str(maxF30))
                if i == 1:
                    minF30 = 2
                    pmaxF30 = maxF30 + 2

                else:
                    pmaxF30 = pmaxF30 + maxF30

                for j in range(minF30, pmaxF30):
                    print("Link z komórki C" + str(j))
                    link = worksheet.cell(column=3, row=j).value
                    if (link == None):
                        print("Nie ma już więcej linków")
                        QApplication.processEvents()
                        status.setStyleSheet('color: rgb(227, 62, 51);')
                        status.setText("Zakończono")
                        bot.close()
                        sys.exit(1)

                    print(link)
                    bot.get(link)
                    sleep(randint(3, 6))
                    # Sprawdzanie czy zaobserwowany
                    try:
                        bot.find_element_by_xpath(wyslaneZaproszenie)
                        worksheet.cell(row=j, column=4, value='Zaobserwowany wcześniej')
                        worksheet.cell(row=j, column=5, value='Konto Prywatne')
                    except:
                        print("Nie Zaobserwowany wcześniej")
                        try:
                            bot.find_element_by_xpath(wyslijWiadomosc)
                            worksheet.cell(row=j, column=4, value='Zaobserwowany wcześniej')
                            worksheet.cell(row=j, column=5, value='Konto Publiczne')
                        except:
                            print("Nie Zaobserwowany wcześniej")
                            try:
                                bot.find_element_by_xpath(obserwuj).click()
                                worksheet.cell(row=j, column=4, value='Zaobserwowany')
                                sleep(5)
                                try:
                                    bot.find_element_by_xpath(wyslijWiadomosc)
                                    follow = follow + 1
                                    worksheet.cell(row=j, column=5, value='Konto Publiczne')
                                    sleep(randint(3, 6))
                                    followed_count.setText(str(follow))
                                    if bot.find_element_by_xpath(wyslijWiadomosc):
                                        # Photo/Like/Com
                                        if (com == True or likeed == True):
                                            try:
                                                bot.find_element_by_xpath('//*[@class="eLAPa"]').click()
                                                sleep(1)
                                                if (likeed == True):
                                                    glc = randint(1, mpc)
                                                    mc = round(glc / 2)
                                                    print("mid " + str(mc))
                                                    lce = 0
                                                    try:
                                                        for k in range(0, glc):
                                                            if (k == 0 or k == mc or k == glc - 1):
                                                                try:
                                                                    bot.find_element_by_xpath(like).click()
                                                                except:
                                                                    bot.find_element_by_xpath(
                                                                        '/html/body/div[6]/div[2]/div/article/div[3]/section[1]/span[1]/button').click()
                                                                sleep(1)
                                                                lc = lc + 1
                                                                lce = lce + 1
                                                                worksheet.cell(row=j, column=6, value=lce)
                                                                liked_count.setText(str(lc))
                                                                if (k == mc):
                                                                    if (com == True):
                                                                        comnumb = randint(0, com_com - 1)
                                                                        comment = com_list[comnumb]
                                                                        print(comment)
                                                                        clipboard.copy(comment)
                                                                        # umieszenie komentarza
                                                                        comment_1 = bot.find_element_by_xpath(koment)
                                                                        comment_1.click()
                                                                        sleep(2)
                                                                        comment_2 = bot.find_element_by_class_name(
                                                                            'Ypffh')
                                                                        comment_2.send_keys(Keys.LEFT_CONTROL, "v")
                                                                        sleep(2)
                                                                        bot.find_element_by_xpath(publish).click()
                                                                        print("Comment")
                                                                        cm = cm + 1
                                                                        comment_count.setText(str(cm))
                                                                        worksheet.cell(row=j, column=7, value=str(mc))
                                                                        worksheet.cell(row=j, column=8, value=comment)
                                                                bot.find_element_by_xpath(nextphoto).click()
                                                                sleep(2)
                                                            else:
                                                                bot.find_element_by_xpath(nextphoto).click()
                                                                sleep(2)
                                                    except:
                                                        print("Nie ma więcej Zdjęć")
                                            except:
                                                print("Photoproblem")
                                                worksheet.cell(row=j, column=6,
                                                               value="Brak Zdjęć")
                                                worksheet.cell(row=j, column=7,value="Brak zdjęcia by umieścić komentarz")

                                except:
                                    print("Konto nie jest publiczne")
                                    follow = follow + 1
                                    followed_count.setText(str(follow))
                                    try:
                                        bot.find_element_by_xpath(wyslaneZaproszenie)
                                        worksheet.cell(row=j, column=5, value='Konto Prywatne')

                                    except:
                                        print("RIP")
                                workbook.save('DataProfile.xlsx')
                            except:
                                print("Problem z linkiem")
                                worksheet.cell(row=j, column=4, value='Problem z Linkiem')
                    finally:
                        workbook.save('DataProfile.xlsx')
                        file.close()
                    slow = randint(2, 4)
                    sleep(slow)
                    print(slowwly)
                    slowwly = slowwly + slow
                minF30 = minF30 + maxF30
                timeslow = 1700 - slowwly
                print("Waiting: " + str(timeslow) + " sec " + str(round(timeslow / 60)) + " min")
                for remaining in range(timeslow, 0, -1):
                    timer.setText(str(remaining) + " s")
                    sleep(1)
                timer.setText("0 s")
        else:
            print("Brak Pliku obok pliku z botem")

        ### Ustawienie Zakończono w Info
        if problem == 0:
            bot.close()
            QApplication.processEvents()
            status.setStyleSheet('color: rgb(227, 62, 51);')
            status.setText("Zakończono")

    ## STWORZONE PRZEZ DANIEL ZIELINSKI Instagram: @Zielu971

    ### Wyślij wiadomośc ###
    def StartMessageSend(self):
        ### Ustawienie Aktywny w Info
        status.setText("Aktywny")
        status.setStyleSheet('color: rgb(102, 137, 37) ')
        QApplication.processEvents()

        log = l.text()
        pas = p.text()

        ##SPRAWDZANIE
        if len(log) == 0 or len(pas) == 0:
            popupTitle = "Insta Login"
            popupContent = "Uzupełni Dane do konta IG"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

        if os.path.exists("HaszProfile.xlsx"):
            self.thread_Message()

        else:
            popupTitle = "Wyślij Wiadomość"
            popupContent = "Utwórz plik o nazwie HaszProfile.xlsx lub pobierz profile z #"
            self.popup(popupTitle, popupContent)
            self.problem_Status()

    def MessageSend(self):
        ### wartości globalne informacji
        global problem_text

        msgsent = 0

        ### Skrypty ###
        search = '//*[@placeholder="Szukaj..."]'
        finded = '//*[@aria-label="Przełącz wybór"]'
        dalej = '/html/body/div[2]/div/div/div[1]/div/div[2]/div/button'
        clk_message = '//*[@placeholder="Wyślij wiadomość..."]'
        send = '//*[text()="Wyślij"]'

        #### KOMENTARZE Z PLIKU ####
        file = open("Wiadomość.txt", encoding="utf-8")
        lines = file.read()

        mes_list = lines.splitlines()
        file.close()
        mes_mes = len(mes_list)

        self.instaLogin()
        sleep(randint(3, 6))

        if os.path.exists('HaszProfile.xlsx'):
            workbook = openpyxl.load_workbook('HaszProfile.xlsx')
            worksheet = workbook.active
            worksheet.cell(row=1, column=4, value='Wysłana wiadomość')
            pCount = len(worksheet['A'])
            workbook.save('HaszProfile.xlsx')
        sleep(randint(2, 4))
        for i in range(1, pCount):
            bot.get('https://www.instagram.com/direct/new/')
            sleep(randint(3, 6))
            bot.find_element_by_xpath(search).click()
            sleep(3)
            profil = worksheet.cell(column=2, row=i + 1).value
            # print(profil)
            bot.find_element_by_xpath(search).send_keys(profil)
            sleep(randint(1, 4))
            try:
                bot.find_element_by_xpath(finded).click()
                sleep(randint(3, 6))
                bot.find_element_by_xpath(dalej).click()
                sleep(randint(3, 6))
                bot.find_element_by_xpath(clk_message).click()
                numMes = randint(0, mes_mes - 1)
                sendmsg = mes_list[numMes]
                clipboard.copy(sendmsg)
                sleep(2)
                bot.find_element_by_xpath(clk_message).send_keys(Keys.LEFT_CONTROL, "v")
                msgsent = msgsent + 1
                sendmsg.setText(str(msgsent))
                sleep(5)
                worksheet.cell(row=i + 1, column=4, value=sendmsg)
                workbook.save('HaszProfile.xlsx')
                try:
                    bot.find_element_by_xpath('/html/body/div[1]/section/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div/div/div[3]/button').click()
                except:
                    bot.find_element_by_xpath(send).click()
                sleep(3)
            except:
                sendmsg = "Nie moge znaleźć profilu"
                worksheet.cell(row=i + 1, column=4, value=sendmsg)
                workbook.save('HaszProfile.xlsx')
        ### END ###
        bot.close()
        QApplication.processEvents()
        status.setStyleSheet('color: rgb(227, 62, 51);')
        status.setText("Zakończono")

    #UNFOLLOW
    def addAccount(self):
            global addacc
            with open("Res/ZapiszFollow.daz", encoding="utf-8") as f:
                    file_text = f.read()
                    f.close()
            self.comedit = QtWidgets.QMainWindow()
            self.ui = ComEdit()
            self.ui.ComEditsetup(self.comedit)
            self.ui.LoginIGmainLabel.setText("Edtuj Obserwowanych")
            self.ui.textEdit.setText(file_text)
            self.comedit.show()
            self.ui.CloseButton.clicked.connect(self.closeSaveAcc)
            self.ui.Website.clicked.connect(self.saveAddAcc)
            addacc = self.ui.textEdit
            self.ui.Website.clicked.connect(self.comedit.close)
    def saveAddAcc(self):
        txt = addacc.toPlainText()
        f = open("Res/ZapiszFollow.daz", "w", encoding="utf-8")
        f.write(txt)
        f.close()
    def closeSaveAcc(self):
        popuptitle = "Edytuj Obserwowanych"
        popupContent = "Czy zapisałeś zmiany ?"
        who = "accclose"
        self.decidePopup(popuptitle, popupContent, who)

    def StartUnfollow(self):
        global ln, log
        ### Ustawienie Aktywny w Info
        status.setText("Aktywny")
        status.setStyleSheet('color: rgb(102, 137, 37) ')
        QApplication.processEvents()

        log = l.text()
        pas = p.text()

        ##SPRAWDZANIE
        if len(log) == 0 or len(pas) == 0:
                popupTitle = "Insta Login"
                popupContent = "Uzupełni Dane do konta IG"
                self.popup(popupTitle, popupContent)
                self.problem_Status()
        else:

                txt = open("Res/ZapiszFollow.daz", "r")
                ln = txt.readlines()
                if len(ln) == 0:
                        print(0)

                else:
                        print(1)
                self.thread_Unfollow()
    def Unfollow(self):
        data = open("Res/ZapiszFollow.daz", encoding="utf-8")
        myFollow = data.readlines()
        print(myFollow)
        slowwly = 0
        unfollowed = 0
        unfollow = unf
        svunf = len(ln)
        Loopmax = 14
        timer = clock
        self.instaLogin()
        bot.get("https://www.instagram.com/" + log +"/")
        bot.find_element_by_xpath("//*[text()='Obserwowani: ']").click()
        sleep(randint(3,6))

        for i in range(1, Loopmax):
                print("Pętla " + str(i) + " z " + str(Loopmax))
                maxF30 = randint(17, 21)
                print("odobserwowanych " + str(maxF30))
                if i == 1:
                        minF30 = 0
                        pmaxF30 = maxF30

                else:
                        pmaxF30 = pmaxF30 + maxF30

                for j in range(minF30, pmaxF30):
                        if j % 4 ==0:
                                fBody = bot.find_element_by_xpath("//div[@class='isgrP']")
                                bot.execute_script('arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].offsetHeight;',fBody)

                        who = bot.execute_script('return document.getElementsByClassName("FPmhX notranslate  _0imsa ")[' + str(j) + '].innerText')
                        print(who)
                        item1 = who
                        item2 = item1 + "\n"

                        if item1 in myFollow:
                                print("jest zapisany "+item1)
                        elif item2 in myFollow:
                                print("jest zapisany "+item2)
                        else:
                              print(j)
                              #Przycisk Odobserwuj
                              try:
                                    bot.execute_script('function getElementByXpath(path) {return document.evaluate(path, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;}'
                                                       'getElementByXpath("/html/body/div[6]/div/div/div[3]/ul/div/li['+str(j+1)+']/div/div[3]/button").click();')
                                    sleep(randint(3,6))
                              except:
                                    print("problem z Odobserwuj 1")
                                    try:
                                        bot.execute_script('function getElementByXpath(path) {return document.evaluate(path, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;}'
                                                       'getElementByXpath("/html/body/div[6]/div/div/div[3]/ul/div/li['+str(j+1)+']/div/div[2]/button").click();')
                                        sleep(randint(3, 6))
                                    except:
                                            print("problem z Odobserwuj 2")

                              #Przycisk Przestań obserwować
                              try:
                                      bot.execute_script(
                                              'var button = document.getElementsByClassName("aOOlW -Cab_   ");'
                                              'button[0].click();')
                                      sleep(1)
                              except:
                                      print('problem z Przestań obserwować 1')
                                      try:
                                        bot.execute_script(
                                              'function getElementByXpath(path) {return document.evaluate(path, document, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;}'
                                              'getElementByXpath("/html/body/div[6]/div/div/div[3]").click();')
                                        sleep(randint(3, 6))
                                      except:
                                        print('problem z Przestań obserwować 2')
                                        try:
                                                bot.find_element_by_xpath('/html/body/div[7]/div/div/div/div[3]/button[1]').click()
                                        except:
                                                print('problem z Przestań obserwować 3')

                              unfollowed = unfollowed + 1
                              unfollow.setText(str(unfollowed))
                        slow = randint(2, 4)
                        sleep(slow)
                        #print(slowwly)
                        slowwly = slowwly + slow
                minF30 = minF30 + maxF30
                timeslow = 1700 - slowwly
                #print("Waiting: " + str(timeslow) + " sec " + str(round(timeslow / 60)) + " min")
                for remaining in range(timeslow, 0, -1):
                        timer.setText(str(remaining) + " s")
                        sleep(1)
                timer.setText("0 s")


    ###  popup i inne

    def perror(self):
        problem_txt = problem_text + "\n"
        if os.path.exists("problem.log"):
            er = open("problem.log", "a")
            er.write(problem_txt)
            er.close()

        else:
            er = open("problem.log", "w")
            er.write(problem_txt)
            er.close()

    def problem(self):
        status.setStyleSheet('color: rgb(255, 0 ,0);')
        status.setText("Problem")
        problem_txt = problem_text + "\n"

        if os.path.exists("problem.log"):
            er = open("problem.log", "a")
            er.write(problem_txt)
            er.close()

        else:
            er = open("problem.log", "w")
            er.write(problem_txt)
            er.close()

    def sleep(self):
        a = randint(3, 6)
        for remaining in range(a, 0, -1):
            clock.setText(str(remaining) + " s")
            sleep(1)

    def problem_Status(self):
        status.setStyleSheet('color: rgb(255, 0 ,0);')
        status.setText("Problem")

    def instaLogin(self):
        global bot, problem_text, popupTitle, popupContent, problem

        ### dane początkowe
        login = l.text()
        passwd = p.text()
        ## przeglądarka
        fimrefox = F.isChecked()
        chrome = C.isChecked()

        optionsc = webdriver.ChromeOptions()
        optionsc.add_experimental_option('excludeSwitches', ['enable-logging'])

        options = Options()
        options.add_argument("--headless")
        options.add_argument('--disable-gpu')
        options.add_argument('--log-level=3')

        ###SPRAWDZAM WYPELNIONE POLA###

        if len(login) == 0 or len(passwd) == 0:
            popupTitle = "Insta Login"
            popupContent = "BRAK DANYCH DO LOGOWANIA"
            self.problem_Status()
            self.popup()
            sys.exit(1)

        if fimrefox == True:
            bot = webdriver.Firefox(executable_path="Driver/geckodriver.exe")
        if chrome == True:
            bot = webdriver.Chrome(executable_path='Driver/chromedriver.exe',chrome_options=optionsc, options=options)


        bot.get('https://www.instagram.com/')

        sleep(randint(3, 6))

        try:
            bot.find_element_by_xpath('/html/body/div[3]/div/div/button[1]').click()
        except:
            problem = 1
            problem_text = "Problem z zakceptowaniem plikow Cookie #instaLogin"
            self.perror()

        sleep(randint(3, 6))

        try:
            username = bot.find_element_by_name('username')
            password = bot.find_element_by_name('password')
            username.clear()
            password.clear()
            username.send_keys(login)
            password.send_keys(passwd)
            password.send_keys(Keys.RETURN)
        except:
            problem = 1
            problem_text = "Problem z Logowaniem #instaLogin"
            self.problem()
            sys.exit(1)
            bot.close()
        sleep(randint(3, 6))
        print("Test")
        sleep(2)
        try:
             bot.find_element_by_xpath("/html/body/div[4]/div/div/button[2]").click()
        except:
                print("problem")
        sleep(2)
        try:
            bot.find_element_by_xpath("//*[text()='Zapisz informacje']").click()
        except:
            try:
                bot.find_element_by_xpath('//*[@id="react-root"]/section/main/div/div/div/section/div/button').click()
            except:
                try:
                    sleep(2)
                    bot.find_element_by_xpath('//*[@id="react-root"]/section/main/div/div/div/div/button').click()
                except:
                    try:
                        bot.find_element_by_xpath("//*[text()='Nie teraz']").click()
                    except:

                        problem = 1
                        problem_text = "Problem z przyciskiem Zapisz dane do logowania #instaLogin"
                        self.problem()
                        sys.exit(1)
                        bot.close()

        sleep(randint(3, 6))
        try:
            bot.find_element_by_xpath("//*[text()='Nie teraz']").click()
        except:
            problem = 1
            problem_text = "Problem z przyciskiem Włącz powiadomienia #instaLogin"
            self.problem()

    def decidePopup(self, popupTitle, PopupContent, who):
        global yes, no
        self.dpop = QtWidgets.QMainWindow()
        self.ui = decPopup()
        self.ui.decSetup(self.dpop)
        self.dpop.show()
        self.ui.dPopupTitle.setText(popupTitle)
        self.ui.dPopupContent.setText(PopupContent)
        if (who == "save"):
            self.ui.yesButton.clicked.connect(self.YesSave)
            self.ui.yesButton.clicked.connect(self.dpop.hide)
            self.ui.noButton.clicked.connect(self.NoSave)
            self.ui.noButton.clicked.connect(self.dpop.hide)
        if (who == "del"):
            self.ui.yesButton.clicked.connect(self.YesDel)
            self.ui.yesButton.clicked.connect(self.dpop.hide)
            self.ui.noButton.clicked.connect(self.NoDel)
            self.ui.noButton.clicked.connect(self.dpop.hide)
        if (who == "comclose"):
            self.ui.yesButton.clicked.connect(self.saveComEdit)
            self.ui.yesButton.clicked.connect(self.dpop.hide)
            self.ui.yesButton.clicked.connect(self.comedit.close)
            self.ui.noButton.clicked.connect(self.dpop.close)
            self.ui.noButton.clicked.connect(self.comedit.close)
        if (who == "mesclose"):
            self.ui.yesButton.clicked.connect(self.saveMesEdit)
            self.ui.yesButton.clicked.connect(self.dpop.hide)
            self.ui.yesButton.clicked.connect(self.comedit.close)
            self.ui.noButton.clicked.connect(self.dpop.close)
            self.ui.noButton.clicked.connect(self.comedit.close)
        if (who == "accclose"):
            self.ui.yesButton.clicked.connect(self.saveAddAcc)
            self.ui.yesButton.clicked.connect(self.dpop.hide)
            self.ui.yesButton.clicked.connect(self.comedit.close)
            self.ui.noButton.clicked.connect(self.dpop.close)
            self.ui.noButton.clicked.connect(self.comedit.close)

    def popup(self, popupTitle, PopupContent):

        self.popp = QtWidgets.QMainWindow()
        self.ui = Popup()
        self.ui.setup(self.popp)
        self.popp.show()
        self.ui.popupTitle.setText(popupTitle)
        self.ui.popupContent.setText(PopupContent)
        self.ui.confirmButton.clicked.connect(self.popp.hide)

    def CloseComEdit(self):
        popuptitle = "Edytuj Komentarz"
        popupContent = "Czy zapisałeś zmiany ?"
        who = "comclose"
        self.decidePopup(popuptitle, popupContent, who)

    def CloseMesEdit(self):
        popuptitle = "Edytuj Wiadomości"
        popupContent = "Czy zapisałeś zmiany ?"
        who = "mesclose"
        self.decidePopup(popuptitle, popupContent, who)

    def comEdit(self):
        global comtext
        with open("Res/Komentarz.daz", encoding="utf-8") as f:
            file_text = f.read()
            f.close()
        self.comedit = QtWidgets.QMainWindow()
        self.ui = ComEdit()
        self.ui.ComEditsetup(self.comedit)
        self.ui.textEdit.setText(file_text)
        self.comedit.show()
        self.ui.CloseButton.clicked.connect(self.CloseComEdit)
        self.ui.Website.clicked.connect(self.saveComEdit)
        comtext = self.ui.textEdit
        self.ui.Website.clicked.connect(self.comedit.close)

    def mesEdit(self):
        global mestext
        with open("Res/Wiadomość.daz", encoding="utf-8") as f:
            file_text = f.read()
            f.close()
        self.comedit = QtWidgets.QMainWindow()
        self.ui = ComEdit()
        self.ui.ComEditsetup(self.comedit)
        self.ui.LoginIGmainLabel.setText("Edtuj Wiadomości")
        self.ui.textEdit.setText(file_text)
        self.comedit.show()
        self.ui.CloseButton.clicked.connect(self.CloseMesEdit)
        self.ui.Website.clicked.connect(self.saveMesEdit)
        mestext = self.ui.textEdit
        self.ui.Website.clicked.connect(self.comedit.close)

    def saveMesEdit(self):
        txt = mestext.toPlainText()
        f = open("Res/Wiadomość.daz", "w", encoding="utf-8")
        f.write(txt)
        f.close()

    def saveComEdit(self):
        txt = comtext.toPlainText()
        f = open("Res/Komentarz.daz", "w", encoding="utf-8")
        f.write(txt)
        f.close()

    def ContactUS(self):
        ## przeglądarka
        fimrefox = F.isChecked()
        chrome = C.isChecked()

        options = webdriver.ChromeOptions()
        options.add_experimental_option('excludeSwitches', ['enable-logging'])


        if fimrefox == True:
            bot = webdriver.Firefox(executable_path="Driver/geckodriver.exe")
        if chrome == True:
            bot = webdriver.Chrome(executable_path='Driver/chromedriver.exe', options=options)


        bot.get('https://szkolainsta.pl/kontakt/')


if __name__ == "__main__":
    import sys, res

    app = QtWidgets.QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())
